#!/usr/bin/env python3
import argparse, base64, json, os
from datetime import date, datetime
from pathlib import Path
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import load_workbook

OUT = Path('data.enc.json')
MAP = {
    '\u7f16\u53f7':'id', '\u5546\u54c1\u7f29\u7565\u56fe':'thumbnail',
    '\u5546\u54c1\u6807\u9898':'title', '\u5206\u7c7b':'category',
    '\u6210\u8272':'condition', '\u5546\u54c1\u4fe1\u606f':'description',
    '\u6210\u672c\u4ef7':'cost', '\u95f2\u9c7c\u552e\u4ef7':'price',
    '\u9884\u4f30\u5229\u6da6':'profit', '\u5229\u6da6\u7387':'profitRate',
    '\u5e93\u5b58\u6570\u91cf':'stock', '\u4e0a\u67b6\u72b6\u6001':'status',
    '\u56fe\u7247\u94fe\u63a5':'imageUrl', '\u95f2\u9c7c\u94fe\u63a5':'xianyuUrl',
    '\u5907\u6ce8':'notes', '\u66f4\u65b0\u65f6\u95f4':'updatedAt',
}

def b64u(raw):
    return base64.urlsafe_b64encode(raw).decode('ascii').rstrip('=')

def find_book(name):
    if name:
        p = Path(name)
        if not p.exists():
            raise FileNotFoundError(p)
        return p
    books = [p for p in Path('.').glob('*.xlsx') if not p.name.startswith('~$')]
    if len(books) != 1:
        raise RuntimeError('Expected exactly one .xlsx workbook or pass --excel')
    return books[0]

def jv(v):
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.isoformat()
    return v

def headers(values):
    out, seen = [], {}
    for i, h in enumerate(values, 1):
        label = str(h).strip() if h not in (None, '') else f'Column {i}'
        key = MAP.get(label, f'col{i}')
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            key = f'{key}{seen[key]}'
        out.append({'key': key, 'label': label})
    return out

def num(v):
    return float(v) if isinstance(v, (int, float)) else None

def compact(rows):
    lr = lc = 0
    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            if v not in (None, ''):
                lr, lc = max(lr, r), max(lc, c)
    return [row[:lc] for row in rows[:lr]]

def matrix(ws):
    return compact([[jv(v) for v in row] for row in ws.iter_rows(values_only=True)])

def inventory(wsv, wsf):
    cols = headers([c.value for c in wsf[1]])
    rows = []
    for r in range(2, wsf.max_row + 1):
        item, has = {}, False
        for c, col in enumerate(cols, 1):
            v = wsv.cell(r, c).value
            item[col['key']] = jv(v)
            has = has or v not in (None, '')
        if not has:
            continue
        cost, price = num(item.get('cost')), num(item.get('price'))
        if item.get('profit') in (None, '') and cost is not None and price is not None:
            item['profit'] = round(price - cost, 4)
        profit = num(item.get('profit'))
        if item.get('profitRate') in (None, '') and profit is not None and price:
            item['profitRate'] = round(profit / price, 6)
        rows.append(item)
    return rows, cols

def read_excel(path):
    wbv = load_workbook(path, data_only=True)
    wbf = load_workbook(path, data_only=False)
    rows, cols = inventory(wbv.worksheets[0], wbf.worksheets[0])
    return {
        'schemaVersion': 1,
        'generatedAt': datetime.now().astimezone().isoformat(timespec='seconds'),
        'sourceWorkbook': path.name,
        'inventoryColumns': cols,
        'inventory': rows,
        'sheets': [{'name': ws.title, 'rows': matrix(ws)} for ws in wbv.worksheets],
    }

def encrypt(payload):
    key, nonce = os.urandom(32), os.urandom(12)
    plain = json.dumps(payload, ensure_ascii=False, separators=(',', ':')).encode('utf-8')
    enc = AESGCM(key).encrypt(nonce, plain, None)
    return {
        'version': '1',
        'algorithm': 'AES-256-GCM',
        'keyFormat': 'base64url-raw-256',
        'nonce': b64u(nonce),
        'ciphertext': b64u(enc),
    }, b64u(key)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel')
    ap.add_argument('--base-url')
    args = ap.parse_args()
    book = find_book(args.excel)
    payload = read_excel(book)
    enc, key = encrypt(payload)
    with OUT.open('w', encoding='utf-8', newline='\n') as fh:
        json.dump(enc, fh, indent=2)
    print(f'Read workbook: {book}')
    print('Inventory rows: {}'.format(len(payload['inventory'])))
    print('Sheets mapped: {}'.format(len(payload['sheets'])))
    print(f'Wrote encrypted data: {OUT}')
    print('\nAccess key, do not commit or paste into README:')
    print(key)
    print('\nShare URL:')
    base = args.base_url.rstrip('/') if args.base_url else 'https://<user>.github.io/<repo>'
    print(base + '/#key=' + key)

if __name__ == '__main__':
    main()
